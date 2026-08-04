import os
import io
import json
import time
import sys
import base64
import smtplib
from email.mime.text import MIMEText
from datetime import datetime, timezone, timedelta

import requests
import openpyxl
from google.oauth2 import service_account
from google.oauth2.credentials import Credentials as UserCredentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaFileUpload

from nacl import encoding, public

# ---------------------------------------------------------------------------
# Configuration (from environment variables / GitHub Secrets)
# ---------------------------------------------------------------------------

IG_USER_ID = os.environ["IG_USER_ID"]
IG_ACCESS_TOKEN = os.environ["IG_ACCESS_TOKEN"]
FB_APP_ID = os.environ["FB_APP_ID"]
FB_APP_SECRET = os.environ["FB_APP_SECRET"]

FB_PAGE_ID = os.environ["FB_PAGE_ID"]
FB_PAGE_ACCESS_TOKEN = os.environ["FB_PAGE_ACCESS_TOKEN"]

YT_CLIENT_ID = os.environ["YT_CLIENT_ID"]
YT_CLIENT_SECRET = os.environ["YT_CLIENT_SECRET"]
YT_REFRESH_TOKEN = os.environ["YT_REFRESH_TOKEN"]

GOOGLE_DRIVE_CREDENTIALS_JSON = os.environ["GOOGLE_DRIVE_CREDENTIALS_JSON"]
DRIVE_FOLDER_ID = os.environ["DRIVE_FOLDER_ID"]  # the Xamo_AutoPoster folder

HEALTHCHECK_PING_URL = os.environ["HEALTHCHECK_PING_URL"]

# --- Alerting ---
GMAIL_ADDRESS = os.environ["GMAIL_ADDRESS"]
GMAIL_APP_PASSWORD = os.environ["GMAIL_APP_PASSWORD"]
ALERT_EMAIL_TO = os.environ.get("ALERT_EMAIL_TO", GMAIL_ADDRESS)

CALLMEBOT_PHONE = os.environ["CALLMEBOT_PHONE"]
CALLMEBOT_APIKEY = os.environ["CALLMEBOT_APIKEY"]

# --- GitHub (for auto-updating the IG token secret after refresh) ---
GH_PAT = os.environ["GH_PAT"]
GH_REPO = os.environ["GITHUB_REPOSITORY"]  # auto-provided by Actions, e.g. "user/xamo-auto-poster"

# YouTube quota safety settings
YT_MIN_HOURS_BETWEEN_SWEEPS = 4  # a "sweep" = one batch of YT uploads

# IG long-lived token: refresh proactively well before the real 60-day expiry
IG_TOKEN_REFRESH_AFTER_DAYS = 50

# A platform gets marked "skipped_error" (and stops being retried) after this many failures.
# NOTE: this 3-strikes rule applies to IG and FB only. YouTube failures during a sweep
# are simply labeled "failed" permanently (see run() for why) and are never retried.
MAX_FAILURES_PER_PLATFORM = 3

QUEUE_FILENAME = "queue.xlsx"
TOKEN_STATE_FILENAME = "token_state.json"
YT_SWEEP_STATE_FILENAME = "yt_sweep_state.json"
VIDEOS_SUBFOLDER = "videos"
POSTED_SUBFOLDER = "posted"

LOCAL_WORKDIR = "/tmp/xamo_poster"
os.makedirs(LOCAL_WORKDIR, exist_ok=True)

GRAPH_API_VERSION = "v21.0"


def log(msg):
    print(f"[{datetime.now(timezone.utc).isoformat()}] {msg}", flush=True)


# ---------------------------------------------------------------------------
# Alerting: email primary, WhatsApp (CallMeBot) failsafe
# ---------------------------------------------------------------------------

def send_email_alert(subject, body):
    msg = MIMEText(body)
    msg["Subject"] = f"[Xamo Auto Poster] {subject}"
    msg["From"] = GMAIL_ADDRESS
    msg["To"] = ALERT_EMAIL_TO

    with smtplib.SMTP_SSL("smtp.gmail.com", 465, timeout=20) as server:
        server.login(GMAIL_ADDRESS, GMAIL_APP_PASSWORD)
        server.sendmail(GMAIL_ADDRESS, [ALERT_EMAIL_TO], msg.as_string())


def send_whatsapp_alert(message):
    resp = requests.get(
        "https://api.callmebot.com/whatsapp.php",
        params={"phone": CALLMEBOT_PHONE, "text": message, "apikey": CALLMEBOT_APIKEY},
        timeout=20,
    )
    resp.raise_for_status()


def send_alert(subject, body):
    """Email first. If email fails for any reason, fall back to WhatsApp so
    you still hear about it. Never lets an alerting failure crash the run."""
    try:
        send_email_alert(subject, body)
        log(f"Alert sent via email: {subject}")
        return
    except Exception as e:
        log(f"Email alert failed ({e}), falling back to WhatsApp...")

    try:
        send_whatsapp_alert(f"{subject}: {body}"[:1000])
        log(f"Alert sent via WhatsApp fallback: {subject}")
    except Exception as e:
        log(f"WhatsApp fallback alert ALSO failed ({e}). No alert could be delivered.")


# ---------------------------------------------------------------------------
# Google Drive helpers
# ---------------------------------------------------------------------------

def get_drive_service():
    creds_dict = json.loads(GOOGLE_DRIVE_CREDENTIALS_JSON)
    creds = service_account.Credentials.from_service_account_info(
        creds_dict, scopes=["https://www.googleapis.com/auth/drive"]
    )
    return build("drive", "v3", credentials=creds)


def find_child_folder(drive, parent_id, name):
    query = (
        f"'{parent_id}' in parents and name = '{name}' "
        "and mimeType = 'application/vnd.google-apps.folder' and trashed = false"
    )
    results = drive.files().list(
        q=query, fields="files(id, name)",
        supportsAllDrives=True, includeItemsFromAllDrives=True,
    ).execute()
    files = results.get("files", [])
    if not files:
        raise RuntimeError(f"Could not find subfolder '{name}' inside folder {parent_id}")
    return files[0]["id"]


def find_file_in_folder(drive, folder_id, name):
    query = f"'{folder_id}' in parents and name = '{name}' and trashed = false"
    results = drive.files().list(
        q=query, fields="files(id, name)",
        supportsAllDrives=True, includeItemsFromAllDrives=True,
    ).execute()
    files = results.get("files", [])
    return files[0]["id"] if files else None


def download_file(drive, file_id, local_path):
    request = drive.files().get_media(fileId=file_id, supportsAllDrives=True)
    with io.FileIO(local_path, "wb") as fh:
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
    return local_path


def upload_or_replace_file(drive, folder_id, filename, local_path, existing_file_id=None):
    media = MediaFileUpload(local_path, resumable=True)
    if existing_file_id:
        drive.files().update(fileId=existing_file_id, media_body=media, supportsAllDrives=True).execute()
        return existing_file_id
    else:
        file_metadata = {"name": filename, "parents": [folder_id]}
        created = drive.files().create(
            body=file_metadata, media_body=media, fields="id", supportsAllDrives=True,
        ).execute()
        return created["id"]


def move_file(drive, file_id, from_folder_id, to_folder_id):
    drive.files().update(
        fileId=file_id,
        addParents=to_folder_id,
        removeParents=from_folder_id,
        fields="id, parents",
        supportsAllDrives=True,
    ).execute()


def load_json_state(drive, folder_id, filename, default):
    file_id = find_file_in_folder(drive, folder_id, filename)
    if not file_id:
        return dict(default), None
    local_path = os.path.join(LOCAL_WORKDIR, filename)
    download_file(drive, file_id, local_path)
    with open(local_path, "r") as f:
        return json.load(f), file_id


def save_json_state(drive, folder_id, filename, data, file_id):
    local_path = os.path.join(LOCAL_WORKDIR, filename)
    with open(local_path, "w") as f:
        json.dump(data, f)
    return upload_or_replace_file(drive, folder_id, filename, local_path, existing_file_id=file_id)


# ---------------------------------------------------------------------------
# Queue (queue.xlsx) helpers
# ---------------------------------------------------------------------------

# NOTE on statuses per platform (IG_Status / FB_Status / YT_Status):
#   None / "" / "not_attempted"  -> never touched yet
#   "failed"                     -> attempted, failed, still eligible for retry
#                                    (IG/FB: retried until 3 fails; YT: never retried
#                                    again since the sweep marker has moved past it)
#   "posted"                     -> succeeded
#   "skipped_error"              -> IG/FB only: hit 3 fails, permanently excluded
COLUMNS = [
    "ID", "VideoFilename", "Caption", "YT_Title", "Hashtags",
    "IG_Status", "FB_Status", "YT_Status",
    "IG_Fails", "FB_Fails", "YT_Fails",
    "IG_PostedAt", "FB_PostedAt", "YT_PostedAt",
]

DONE_STATES = ("posted", "skipped_error")
NOT_ATTEMPTED_STATES = (None, "", "not_attempted")


def col_index(name):
    return COLUMNS.index(name) + 1  # openpyxl is 1-indexed


def load_queue(local_path):
    wb = openpyxl.load_workbook(local_path)
    ws = wb.active
    return wb, ws


def get_row_dict(ws, row_num):
    return {col: ws.cell(row=row_num, column=col_index(col)).value for col in COLUMNS}


def set_cell(ws, row_num, col_name, value):
    ws.cell(row=row_num, column=col_index(col_name), value=value)


def all_data_rows(ws):
    """Yields (row_num, row_dict) for every row that actually has a video."""
    for row_num in range(2, ws.max_row + 1):
        row = get_row_dict(ws, row_num)
        if row["VideoFilename"] is not None:
            yield row_num, row


def row_fully_posted(row):
    return row["IG_Status"] == "posted" and row["FB_Status"] == "posted" and row["YT_Status"] == "posted"


# ---------------------------------------------------------------------------
# IG long-lived token auto-refresh + GitHub secret update
# ---------------------------------------------------------------------------

def refresh_ig_token(current_token):
    """Exchange the current token for a fresh 60-day one, using the
    correct endpoint for Instagram-Login-issued tokens (IGAA...)."""
    resp = requests.get(
        "https://graph.instagram.com/refresh_access_token",
        params={
            "grant_type": "ig_refresh_token",
            "access_token": current_token,
        },
        timeout=30,
    ).json()

    if "access_token" not in resp:
        raise RuntimeError(f"IG token refresh failed: {resp}")

    return resp["access_token"]


def encrypt_secret_for_github(public_key_b64, secret_value):
    pk = public.PublicKey(public_key_b64.encode("utf-8"), encoding.Base64Encoder())
    sealed_box = public.SealedBox(pk)
    encrypted = sealed_box.encrypt(secret_value.encode("utf-8"))
    return base64.b64encode(encrypted).decode("utf-8")


def update_github_secret(secret_name, secret_value):
    headers = {
        "Authorization": f"Bearer {GH_PAT}",
        "Accept": "application/vnd.github+json",
    }
    key_resp = requests.get(
        f"https://api.github.com/repos/{GH_REPO}/actions/secrets/public-key",
        headers=headers, timeout=20,
    ).json()
    if "key" not in key_resp:
        raise RuntimeError(f"Could not fetch GitHub public key: {key_resp}")

    encrypted_value = encrypt_secret_for_github(key_resp["key"], secret_value)

    put_resp = requests.put(
        f"https://api.github.com/repos/{GH_REPO}/actions/secrets/{secret_name}",
        headers=headers,
        json={"encrypted_value": encrypted_value, "key_id": key_resp["key_id"]},
        timeout=20,
    )
    if put_resp.status_code not in (201, 204):
        raise RuntimeError(f"Failed to update GitHub secret {secret_name}: {put_resp.status_code} {put_resp.text}")


def maybe_refresh_ig_token(drive):
    """Refreshes the IG token if it's due, updates the GitHub secret, and
    alerts (email -> WhatsApp fallback) if the refresh itself fails.
    Never crashes the run - if refresh fails, we just keep using the old
    token and try again next run."""
    state, state_file_id = load_json_state(
        drive, DRIVE_FOLDER_ID, TOKEN_STATE_FILENAME, {"ig_last_refreshed": None}
    )

    last_refreshed = state.get("ig_last_refreshed")
    due = True
    if last_refreshed:
        last_dt = datetime.fromisoformat(last_refreshed)
        due = (datetime.now(timezone.utc) - last_dt) >= timedelta(days=IG_TOKEN_REFRESH_AFTER_DAYS)

    if not due:
        return IG_ACCESS_TOKEN

    log("IG token refresh is due - attempting refresh...")
    try:
        new_token = refresh_ig_token(IG_ACCESS_TOKEN)
        update_github_secret("IG_ACCESS_TOKEN", new_token)
        state["ig_last_refreshed"] = datetime.now(timezone.utc).isoformat()
        save_json_state(drive, DRIVE_FOLDER_ID, TOKEN_STATE_FILENAME, state, state_file_id)
        log("IG token refreshed successfully and GitHub secret updated.")
        return new_token
    except Exception as e:
        log(f"IG TOKEN REFRESH FAILED: {e}")
        send_alert(
            "IG token refresh FAILED",
            f"The automated Instagram token refresh failed: {e}\n\n"
            f"The bot will keep using the existing token for now, but it may "
            f"expire soon. Please refresh IG_ACCESS_TOKEN manually if this "
            f"keeps happening.",
        )
        return IG_ACCESS_TOKEN  # keep using the old one, don't crash the run


# ---------------------------------------------------------------------------
# Instagram posting
# ---------------------------------------------------------------------------

def upload_temp_github_release_asset(local_video_path, filename):
    """Uploads the video as an asset on a dedicated 'video-hosting' GitHub
    Release, and returns its public download URL. Creates the release if
    it doesn't already exist."""
    headers = {"Authorization": f"Bearer {GH_PAT}", "Accept": "application/vnd.github+json"}

    releases_resp = requests.get(
        f"https://api.github.com/repos/{GH_REPO}/releases/tags/video-hosting",
        headers=headers, timeout=20,
    )
    if releases_resp.status_code == 200:
        release = releases_resp.json()
    else:
        create_resp = requests.post(
            f"https://api.github.com/repos/{GH_REPO}/releases",
            headers=headers,
            json={"tag_name": "video-hosting", "name": "Temporary video hosting", "draft": False, "prerelease": True},
            timeout=20,
        )
        release = create_resp.json()

    upload_url = release["upload_url"].split("{")[0]

    # If a stale asset with this same filename is already sitting on the
    # release (e.g. left over from a previous run that crashed before its
    # `finally` cleanup ran), GitHub will reject the new upload with a 422
    # "already_exists" error. Clear it out first so this run isn't blocked.
    for existing_asset in release.get("assets", []):
        if existing_asset.get("name") == filename:
            log(f"IG: found stale leftover asset '{filename}' from a previous run, deleting it first.")
            requests.delete(
                f"https://api.github.com/repos/{GH_REPO}/releases/assets/{existing_asset['id']}",
                headers=headers, timeout=20,
            )

    with open(local_video_path, "rb") as f:
        upload_resp = requests.post(
            f"{upload_url}?name={filename}",
            headers={**headers, "Content-Type": "video/mp4"},
            data=f,
            timeout=120,
        )
    upload_resp.raise_for_status()
    asset = upload_resp.json()
    return asset["browser_download_url"], asset["id"]


def delete_github_release_asset(asset_id):
    headers = {"Authorization": f"Bearer {GH_PAT}", "Accept": "application/vnd.github+json"}
    requests.delete(
        f"https://api.github.com/repos/{GH_REPO}/releases/assets/{asset_id}",
        headers=headers, timeout=20,
    )


def post_to_instagram(video_local_path, caption, access_token):
    base_url = "https://graph.instagram.com"

    video_url, asset_id = upload_temp_github_release_asset(video_local_path, os.path.basename(video_local_path))
    log(f"IG: temporarily hosted video at {video_url}")

    try:
        create_resp = requests.post(
            f"{base_url}/{IG_USER_ID}/media",
            data={
                "media_type": "REELS",
                "video_url": video_url,
                "caption": caption,
                "access_token": access_token,
            },
        ).json()

        if "id" not in create_resp:
            log(f"IG: failed to create container: {create_resp}")
            return False

        container_id = create_resp["id"]

        status = None
        for _ in range(30):
            time.sleep(10)
            status_resp = requests.get(
                f"{base_url}/{container_id}",
                params={"fields": "status_code", "access_token": access_token},
            ).json()
            status = status_resp.get("status_code")
            if status == "FINISHED":
                break
            if status == "ERROR":
                log(f"IG: container processing failed: {status_resp}")
                return False

        if status != "FINISHED":
            log(f"IG: container did not finish processing in time (last status: {status})")
            return False

        publish_resp = requests.post(
            f"{base_url}/{IG_USER_ID}/media_publish",
            data={"creation_id": container_id, "access_token": access_token},
        ).json()

        if "id" not in publish_resp:
            log(f"IG: publish failed: {publish_resp}")
            return False

        log(f"IG: published successfully, media id {publish_resp['id']}")
        return True

    finally:
        delete_github_release_asset(asset_id)
        log("IG: cleaned up temporary hosted video.")


# ---------------------------------------------------------------------------
# Facebook Page posting
# ---------------------------------------------------------------------------

def post_to_facebook_page(video_local_path, caption):
    base_url = f"https://graph-video.facebook.com/{GRAPH_API_VERSION}"

    with open(video_local_path, "rb") as f:
        files = {"source": f}
        data = {"description": caption, "access_token": FB_PAGE_ACCESS_TOKEN}
        resp = requests.post(f"{base_url}/{FB_PAGE_ID}/videos", data=data, files=files).json()

    if "id" not in resp:
        log(f"FB: post failed: {resp}")
        return False

    log(f"FB: posted successfully, video id {resp['id']}")
    return True


# ---------------------------------------------------------------------------
# YouTube posting
# ---------------------------------------------------------------------------

def get_youtube_service():
    creds = UserCredentials(
        token=None,
        refresh_token=YT_REFRESH_TOKEN,
        client_id=YT_CLIENT_ID,
        client_secret=YT_CLIENT_SECRET,
        token_uri="https://oauth2.googleapis.com/token",
        scopes=["https://www.googleapis.com/auth/youtube.upload"],
    )
    return build("youtube", "v3", credentials=creds)


def post_to_youtube(video_local_path, yt_title, caption, publish_at_iso):
    """publish_at_iso: RFC3339 UTC timestamp (e.g. '2026-07-26T14:00:00Z').
    Uploads now, but the video is scheduled (privacyStatus='private' +
    publishAt) to go public at that time -- this is how we stagger visibility
    for videos that are all actually uploaded in the same sweep."""
    try:
        youtube = get_youtube_service()
        body = {
            "snippet": {
                "title": yt_title[:100],
                "description": caption,
                "categoryId": "22",
            },
            "status": {
                "privacyStatus": "private",
                "publishAt": publish_at_iso,
                "selfDeclaredMadeForKids": False,
            },
        }
        media = MediaFileUpload(video_local_path, chunksize=-1, resumable=True, mimetype="video/mp4")
        request = youtube.videos().insert(part="snippet,status", body=body, media_body=media)
        response = None
        while response is None:
            status, response = request.next_chunk()
        log(f"YT: uploaded successfully, video id {response['id']}, scheduled for {publish_at_iso}")
        return True
    except Exception as e:
        log(f"YT: upload failed: {e}")
        return False


# ---------------------------------------------------------------------------
# Healthcheck ping
# ---------------------------------------------------------------------------

def ping_healthcheck(success=True, message=""):
    try:
        url = HEALTHCHECK_PING_URL if success else f"{HEALTHCHECK_PING_URL}/fail"
        requests.post(url, data=message.encode("utf-8"), timeout=10)
    except Exception as e:
        log(f"Healthcheck ping failed (non-fatal): {e}")


# ---------------------------------------------------------------------------
# Per-platform attempt helper (IG / FB: retry counting + skip-after-N)
# ---------------------------------------------------------------------------

def attempt_platform_with_retries(ws, row_num, row, status_col, fails_col, posted_at_col,
                                   platform_label, post_fn):
    """Used for IG and FB. Runs post_fn() if this platform isn't already
    done. Updates status, failure count, and posted-at timestamp. Sends an
    alert + marks skipped_error if MAX_FAILURES_PER_PLATFORM is reached."""
    if row[status_col] in DONE_STATES:
        return False  # nothing to do, not attempted

    log(f"Attempting {platform_label} post for row {row_num}...")
    success = post_fn()

    if success:
        set_cell(ws, row_num, status_col, "posted")
        set_cell(ws, row_num, posted_at_col, datetime.now(timezone.utc).isoformat())
        set_cell(ws, row_num, fails_col, 0)
    else:
        fails = (row[fails_col] or 0) + 1
        set_cell(ws, row_num, fails_col, fails)
        if fails >= MAX_FAILURES_PER_PLATFORM:
            set_cell(ws, row_num, status_col, "skipped_error")
            log(f"{platform_label}: hit {fails} failures on row {row_num}, marking skipped_error.")
            send_alert(
                f"{platform_label} post skipped after {fails} failures",
                f"Row {row_num} ({row['VideoFilename']}) failed to post to "
                f"{platform_label} {fails} times in a row and has been marked "
                f"'skipped_error'. It will NOT be retried automatically. "
                f"Check the row in queue.xlsx and fix/reset it manually.",
            )
        else:
            set_cell(ws, row_num, status_col, "failed")

    return True  # was attempted


def attempt_youtube_no_retry(ws, row_num, row, post_fn):
    """Used for YT sweeps only. No fail-count tracking, no skipped_error --
    on failure we just label 'failed' and move on for good (per design, the
    sweep marker moves past this row regardless and it will never be
    revisited)."""
    if row["YT_Status"] in DONE_STATES:
        return False

    log(f"Attempting YouTube post for row {row_num}...")
    success = post_fn()

    if success:
        set_cell(ws, row_num, "YT_Status", "posted")
        set_cell(ws, row_num, "YT_PostedAt", datetime.now(timezone.utc).isoformat())
    else:
        set_cell(ws, row_num, "YT_Status", "failed")
        log(f"YouTube: row {row_num} failed this sweep -- will not be retried (marker moves past it).")

    return True


def find_video_anywhere(drive, videos_folder_id, posted_folder_id, filename):
    """Looks in /videos first, then /posted, since IG+FB may have already
    moved the file by the time YT's sweep gets to it (shouldn't normally
    happen now that moves only occur once ALL three platforms are posted,
    but kept as a safety net)."""
    file_id = find_file_in_folder(drive, videos_folder_id, filename)
    if file_id:
        return file_id
    return find_file_in_folder(drive, posted_folder_id, filename)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def run():
    log("Starting poster run.")
    drive = get_drive_service()

    ig_access_token = maybe_refresh_ig_token(drive)

    videos_folder_id = find_child_folder(drive, DRIVE_FOLDER_ID, VIDEOS_SUBFOLDER)
    posted_folder_id = find_child_folder(drive, DRIVE_FOLDER_ID, POSTED_SUBFOLDER)

    queue_file_id = find_file_in_folder(drive, DRIVE_FOLDER_ID, QUEUE_FILENAME)
    if not queue_file_id:
        log(f"ERROR: could not find {QUEUE_FILENAME} in the Drive folder.")
        ping_healthcheck(success=False, message="queue.xlsx not found")
        sys.exit(1)

    local_queue_path = os.path.join(LOCAL_WORKDIR, QUEUE_FILENAME)
    download_file(drive, queue_file_id, local_queue_path)
    wb, ws = load_queue(local_queue_path)

    any_attempted = False

    # =======================================================================
    # LANE 1: IG + FB
    #   - "current row" = the FIRST row that has never been attempted at all
    #     on EITHER IG or FB (i.e. both are in NOT_ATTEMPTED_STATES).
    #   - Rows after the current row are not touched this run, period.
    #   - After processing the current row, go back and retry every row
    #     BEFORE it that's still pending (status == "failed", not yet
    #     skipped_error, not yet posted) on IG and/or FB.
    # =======================================================================
    current_row_num = None
    for row_num, row in all_data_rows(ws):
        if row["IG_Status"] in NOT_ATTEMPTED_STATES and row["FB_Status"] in NOT_ATTEMPTED_STATES:
            current_row_num = row_num
            break

    if current_row_num is None:
        log("No untouched row found for IG/FB -- nothing new to start this run.")
    else:
        log(f"Current row for this run (IG/FB): {current_row_num}")

    # Backlog: earlier rows still pending on IG and/or FB (status == "failed").
    backlog_row_nums = []
    if current_row_num is not None:
        for row_num, row in all_data_rows(ws):
            if row_num >= current_row_num:
                break
            if row["IG_Status"] not in DONE_STATES or row["FB_Status"] not in DONE_STATES:
                backlog_row_nums.append(row_num)

    ig_fb_row_nums = ([current_row_num] if current_row_num is not None else []) + backlog_row_nums
    log(f"IG/FB rows to process this run (current + backlog): {ig_fb_row_nums}")

    for row_num in ig_fb_row_nums:
        row = get_row_dict(ws, row_num)
        video_filename = row["VideoFilename"]
        caption = row["Caption"] or ""

        video_file_id = find_file_in_folder(drive, videos_folder_id, video_filename)
        if not video_file_id:
            log(f"ERROR: video file '{video_filename}' not found in /videos. Skipping row {row_num}.")
            continue

        local_video_path = os.path.join(LOCAL_WORKDIR, video_filename)
        download_file(drive, video_file_id, local_video_path)

        any_attempted |= attempt_platform_with_retries(
            ws, row_num, row, "IG_Status", "IG_Fails", "IG_PostedAt", "Instagram",
            lambda: post_to_instagram(local_video_path, caption, ig_access_token),
        )
        any_attempted |= attempt_platform_with_retries(
            ws, row_num, row, "FB_Status", "FB_Fails", "FB_PostedAt", "Facebook",
            lambda: post_to_facebook_page(local_video_path, caption),
        )

    # =======================================================================
    # LANE 2: YouTube -- purely time-gated, per-row independent sweep.
    #   - The ONLY gate is the 4-hour timer (yt_last_swept_at). Nothing
    #     about any other row's IG/FB status can block or delay a row.
    #   - When due, we sweep EVERY row where IG_Status=="posted" and
    #     FB_Status=="posted" and YT_Status isn't already done -- regardless
    #     of position, regardless of whether some earlier row is stuck in
    #     "failed" or "skipped_error". A row that isn't IG+FB posted yet
    #     simply isn't eligible THIS sweep; it will be picked up on
    #     whichever future sweep it becomes eligible on, with no risk of
    #     being permanently skipped over.
    #   - Each eligible row gets one attempt (attempt_youtube_no_retry) --
    #     on failure it's labeled "failed" and not retried later (YT itself
    #     has no strike-counter, per design).
    #   - Videos are all uploaded in this same run; visibility is staggered
    #     via YouTube's publishAt (now, +1h, +2h, +3h, ... in row order).
    # =======================================================================
    sweep_state, sweep_state_file_id = load_json_state(
        drive, DRIVE_FOLDER_ID, YT_SWEEP_STATE_FILENAME,
        {"yt_last_swept_at": None},
    )

    last_swept_at = sweep_state.get("yt_last_swept_at")

    sweep_due = True
    if last_swept_at:
        last_dt = datetime.fromisoformat(last_swept_at)
        sweep_due = (datetime.now(timezone.utc) - last_dt) >= timedelta(hours=YT_MIN_HOURS_BETWEEN_SWEEPS)

    if not sweep_due:
        log("YT sweep not due yet (last sweep < 4h ago) -- skipping YT this run.")
    elif current_row_num is None:
        log("No current IG/FB row yet -- nothing to sweep for YT.")
    else:
        yt_sweep_row_nums = []
        for row_num, row in all_data_rows(ws):
            if row_num > current_row_num:
                break
            if row["YT_Status"] not in DONE_STATES:
                yt_sweep_row_nums.append(row_num)

        log(f"YT sweep is due. Rows in this sweep: {yt_sweep_row_nums}")
        
        now = datetime.now(timezone.utc)
        for offset, row_num in enumerate(yt_sweep_row_nums):
            row = get_row_dict(ws, row_num)
            video_filename = row["VideoFilename"]
            yt_title = row["YT_Title"] or video_filename
            caption = row["Caption"] or ""

            video_file_id = find_video_anywhere(drive, videos_folder_id, posted_folder_id, video_filename)
            if not video_file_id:
                log(f"ERROR: video file '{video_filename}' not found anywhere for YT. Skipping row {row_num}.")
                continue

            local_video_path = os.path.join(LOCAL_WORKDIR, video_filename)
            download_file(drive, video_file_id, local_video_path)

            publish_at = now + timedelta(hours=offset)
            publish_at_iso = publish_at.strftime("%Y-%m-%dT%H:%M:%SZ")

            attempted = attempt_youtube_no_retry(
                ws, row_num, row,
                lambda: post_to_youtube(local_video_path, yt_title, caption, publish_at_iso),
            )
            if attempted:
                any_attempted = True

        # Only the timestamp needs to persist now -- eligibility is always
        # recomputed fresh from each row's own status, so there's no marker
        # row to save, and therefore nothing that can strand a row.
        sweep_state["yt_last_swept_at"] = datetime.now(timezone.utc).isoformat()
        save_json_state(drive, DRIVE_FOLDER_ID, YT_SWEEP_STATE_FILENAME, sweep_state, sweep_state_file_id)

    # =======================================================================
    # Completion sweep: any row now fully posted on IG+FB+YT gets its video
    # moved from /videos to /posted.
    # =======================================================================
    for row_num, row in all_data_rows(ws):
        if row_fully_posted(row):
            video_filename = row["VideoFilename"]
            video_file_id = find_file_in_folder(drive, videos_folder_id, video_filename)
            if video_file_id:
                log(f"Row {row_num}: fully posted on IG+FB+YT -- moving '{video_filename}' to /posted.")
                move_file(drive, video_file_id, videos_folder_id, posted_folder_id)

    wb.save(local_queue_path)
    upload_or_replace_file(drive, DRIVE_FOLDER_ID, QUEUE_FILENAME, local_queue_path, existing_file_id=queue_file_id)

    if any_attempted:
        ping_healthcheck(success=True, message="Run completed with activity")
    else:
        ping_healthcheck(success=True, message="Nothing needed posting this run")

    log("Run complete.")


def main():
    """Top-level wrapper: guarantees that ANY unhandled error still pings
    healthchecks.io as failed and sends an alert, instead of silently dying
    with no trace (which is exactly what would break the keep-alive commit
    and leave you finding out weeks late)."""
    try:
        run()
    except Exception as e:
        log(f"FATAL ERROR in run(): {e}")
        ping_healthcheck(success=False, message=f"Fatal error: {e}")
        try:
            send_alert("Xamo Auto Poster run CRASHED", f"The bot run failed with an unhandled error:\n\n{e}")
        except Exception:
            log("Could not send crash alert either.")
        sys.exit(1)


if __name__ == "__main__":
    main()



# import os
# import io
# import json
# import time
# import sys
# import base64
# import shutil
# import smtplib
# import subprocess
# from email.mime.text import MIMEText
# from datetime import datetime, timezone, timedelta

# import requests
# import openpyxl
# from google.oauth2 import service_account
# from google.oauth2.credentials import Credentials as UserCredentials
# from googleapiclient.discovery import build
# from googleapiclient.http import MediaIoBaseDownload, MediaFileUpload

# from nacl import encoding, public

# try:
#     import yt_dlp
# except ImportError:
#     print("ERROR: yt-dlp is not installed. Add it to requirements.txt.")
#     sys.exit(1)

# # ---------------------------------------------------------------------------
# # Configuration (from environment variables / GitHub Secrets)
# # ---------------------------------------------------------------------------

# IG_USER_ID = os.environ["IG_USER_ID"]
# IG_ACCESS_TOKEN = os.environ["IG_ACCESS_TOKEN"]
# FB_APP_ID = os.environ["FB_APP_ID"]
# FB_APP_SECRET = os.environ["FB_APP_SECRET"]

# FB_PAGE_ID = os.environ["FB_PAGE_ID"]
# FB_PAGE_ACCESS_TOKEN = os.environ["FB_PAGE_ACCESS_TOKEN"]

# YT_CLIENT_ID = os.environ["YT_CLIENT_ID"]
# YT_CLIENT_SECRET = os.environ["YT_CLIENT_SECRET"]
# YT_REFRESH_TOKEN = os.environ["YT_REFRESH_TOKEN"]

# GOOGLE_DRIVE_CREDENTIALS_JSON = os.environ["GOOGLE_DRIVE_CREDENTIALS_JSON"]
# DRIVE_FOLDER_ID = os.environ["DRIVE_FOLDER_ID"]

# HEALTHCHECK_PING_URL = os.environ["HEALTHCHECK_PING_URL"]

# GMAIL_ADDRESS = os.environ["GMAIL_ADDRESS"]
# GMAIL_APP_PASSWORD = os.environ["GMAIL_APP_PASSWORD"]
# ALERT_EMAIL_TO = os.environ.get("ALERT_EMAIL_TO", GMAIL_ADDRESS)

# CALLMEBOT_PHONE = os.environ["CALLMEBOT_PHONE"]
# CALLMEBOT_APIKEY = os.environ["CALLMEBOT_APIKEY"]

# GH_PAT = os.environ["GH_PAT"]
# GH_REPO = os.environ["GITHUB_REPOSITORY"]

# # A platform gets marked "skipped_error" after this many consecutive failures.
# # Applies to Download, IG, and FB. (YT has no strike-counter -- see run().)
# MAX_FAILURES_PER_PLATFORM = 3

# # IG long-lived token: refresh proactively well before the real 60-day expiry
# IG_TOKEN_REFRESH_AFTER_DAYS = 50

# # Every Nth successful run also gets a YouTube upload for that run's row.
# YT_EVERY_N_RUNS = 4

# # Used as the caption/description whenever a downloaded video has no
# # description of its own.
# DEFAULT_CAPTION = "Follow for more #viral #reels #shorts"

# QUEUE_FILENAME = "queue.xlsx"
# TOKEN_STATE_FILENAME = "token_state.json"
# RUN_STATE_FILENAME = "run_state.json"

# LOCAL_WORKDIR = "/tmp/xamo_poster"
# os.makedirs(LOCAL_WORKDIR, exist_ok=True)

# GRAPH_API_VERSION = "v21.0"

# PAD_TARGET_W = 1080
# PAD_TARGET_H = 1920


# def log(msg):
#     print(f"[{datetime.now(timezone.utc).isoformat()}] {msg}", flush=True)


# # ---------------------------------------------------------------------------
# # Alerting: email primary, WhatsApp (CallMeBot) failsafe
# # ---------------------------------------------------------------------------

# def send_email_alert(subject, body):
#     msg = MIMEText(body)
#     msg["Subject"] = f"[Xamo Auto Poster] {subject}"
#     msg["From"] = GMAIL_ADDRESS
#     msg["To"] = ALERT_EMAIL_TO

#     with smtplib.SMTP_SSL("smtp.gmail.com", 465, timeout=20) as server:
#         server.login(GMAIL_ADDRESS, GMAIL_APP_PASSWORD)
#         server.sendmail(GMAIL_ADDRESS, [ALERT_EMAIL_TO], msg.as_string())


# def send_whatsapp_alert(message):
#     resp = requests.get(
#         "https://api.callmebot.com/whatsapp.php",
#         params={"phone": CALLMEBOT_PHONE, "text": message, "apikey": CALLMEBOT_APIKEY},
#         timeout=20,
#     )
#     resp.raise_for_status()


# def send_alert(subject, body):
#     """Email first. If email fails for any reason, fall back to WhatsApp so
#     you still hear about it. Never lets an alerting failure crash the run."""
#     try:
#         send_email_alert(subject, body)
#         log(f"Alert sent via email: {subject}")
#         return
#     except Exception as e:
#         log(f"Email alert failed ({e}), falling back to WhatsApp...")

#     try:
#         send_whatsapp_alert(f"{subject}: {body}"[:1000])
#         log(f"Alert sent via WhatsApp fallback: {subject}")
#     except Exception as e:
#         log(f"WhatsApp fallback alert ALSO failed ({e}). No alert could be delivered.")


# # ---------------------------------------------------------------------------
# # Google Drive helpers
# # ---------------------------------------------------------------------------

# def get_drive_service():
#     creds_dict = json.loads(GOOGLE_DRIVE_CREDENTIALS_JSON)
#     creds = service_account.Credentials.from_service_account_info(
#         creds_dict, scopes=["https://www.googleapis.com/auth/drive"]
#     )
#     return build("drive", "v3", credentials=creds)


# def find_file_in_folder(drive, folder_id, name):
#     query = f"'{folder_id}' in parents and name = '{name}' and trashed = false"
#     results = drive.files().list(
#         q=query, fields="files(id, name)",
#         supportsAllDrives=True, includeItemsFromAllDrives=True,
#     ).execute()
#     files = results.get("files", [])
#     return files[0]["id"] if files else None


# def download_file(drive, file_id, local_path):
#     request = drive.files().get_media(fileId=file_id, supportsAllDrives=True)
#     with io.FileIO(local_path, "wb") as fh:
#         downloader = MediaIoBaseDownload(fh, request)
#         done = False
#         while not done:
#             _, done = downloader.next_chunk()
#     return local_path


# def upload_or_replace_file(drive, folder_id, filename, local_path, existing_file_id=None):
#     media = MediaFileUpload(local_path, resumable=True)
#     if existing_file_id:
#         drive.files().update(fileId=existing_file_id, media_body=media, supportsAllDrives=True).execute()
#         return existing_file_id
#     else:
#         file_metadata = {"name": filename, "parents": [folder_id]}
#         created = drive.files().create(
#             body=file_metadata, media_body=media, fields="id", supportsAllDrives=True,
#         ).execute()
#         return created["id"]


# def load_json_state(drive, folder_id, filename, default):
#     file_id = find_file_in_folder(drive, folder_id, filename)
#     if not file_id:
#         return dict(default), None
#     local_path = os.path.join(LOCAL_WORKDIR, filename)
#     download_file(drive, file_id, local_path)
#     with open(local_path, "r") as f:
#         return json.load(f), file_id


# def save_json_state(drive, folder_id, filename, data, file_id):
#     local_path = os.path.join(LOCAL_WORKDIR, filename)
#     with open(local_path, "w") as f:
#         json.dump(data, f)
#     return upload_or_replace_file(drive, folder_id, filename, local_path, existing_file_id=file_id)


# # ---------------------------------------------------------------------------
# # Queue (queue.xlsx) helpers
# # ---------------------------------------------------------------------------

# # Columns the human provides:  ID, Link
# # Everything else is written BY the bot as it works.
# COLUMNS = [
#     "ID", "Link", "Caption",
#     "DL_Status", "DL_Fails",
#     "IG_Status", "FB_Status", "YT_Status",
#     "IG_Fails", "FB_Fails",
#     "IG_PostedAt", "FB_PostedAt", "YT_PostedAt",
# ]

# DONE_STATES = ("posted", "skipped_error")
# NOT_ATTEMPTED_STATES = (None, "", "not_attempted")


# def col_index(name):
#     return COLUMNS.index(name) + 1


# def load_queue(local_path):
#     wb = openpyxl.load_workbook(local_path)
#     ws = wb.active
#     return wb, ws


# def get_row_dict(ws, row_num):
#     return {col: ws.cell(row=row_num, column=col_index(col)).value for col in COLUMNS}


# def set_cell(ws, row_num, col_name, value):
#     ws.cell(row=row_num, column=col_index(col_name), value=value)


# def all_data_rows(ws):
#     """Yields (row_num, row_dict) for every row that actually has a link."""
#     for row_num in range(2, ws.max_row + 1):
#         row = get_row_dict(ws, row_num)
#         if row["Link"] is not None:
#             yield row_num, row


# def row_fully_posted(row):
#     return row["IG_Status"] == "posted" and row["FB_Status"] == "posted" and row["YT_Status"] == "posted"


# # ---------------------------------------------------------------------------
# # IG long-lived token auto-refresh + GitHub secret update
# # ---------------------------------------------------------------------------

# def refresh_ig_token(current_token):
#     resp = requests.get(
#         "https://graph.instagram.com/refresh_access_token",
#         params={"grant_type": "ig_refresh_token", "access_token": current_token},
#         timeout=30,
#     ).json()
#     if "access_token" not in resp:
#         raise RuntimeError(f"IG token refresh failed: {resp}")
#     return resp["access_token"]


# def encrypt_secret_for_github(public_key_b64, secret_value):
#     pk = public.PublicKey(public_key_b64.encode("utf-8"), encoding.Base64Encoder())
#     sealed_box = public.SealedBox(pk)
#     encrypted = sealed_box.encrypt(secret_value.encode("utf-8"))
#     return base64.b64encode(encrypted).decode("utf-8")


# def update_github_secret(secret_name, secret_value):
#     headers = {"Authorization": f"Bearer {GH_PAT}", "Accept": "application/vnd.github+json"}
#     key_resp = requests.get(
#         f"https://api.github.com/repos/{GH_REPO}/actions/secrets/public-key",
#         headers=headers, timeout=20,
#     ).json()
#     if "key" not in key_resp:
#         raise RuntimeError(f"Could not fetch GitHub public key: {key_resp}")

#     encrypted_value = encrypt_secret_for_github(key_resp["key"], secret_value)

#     put_resp = requests.put(
#         f"https://api.github.com/repos/{GH_REPO}/actions/secrets/{secret_name}",
#         headers=headers,
#         json={"encrypted_value": encrypted_value, "key_id": key_resp["key_id"]},
#         timeout=20,
#     )
#     if put_resp.status_code not in (201, 204):
#         raise RuntimeError(f"Failed to update GitHub secret {secret_name}: {put_resp.status_code} {put_resp.text}")


# def maybe_refresh_ig_token(drive):
#     state, state_file_id = load_json_state(
#         drive, DRIVE_FOLDER_ID, TOKEN_STATE_FILENAME, {"ig_last_refreshed": None}
#     )
#     last_refreshed = state.get("ig_last_refreshed")
#     due = True
#     if last_refreshed:
#         last_dt = datetime.fromisoformat(last_refreshed)
#         due = (datetime.now(timezone.utc) - last_dt) >= timedelta(days=IG_TOKEN_REFRESH_AFTER_DAYS)

#     if not due:
#         return IG_ACCESS_TOKEN

#     log("IG token refresh is due - attempting refresh...")
#     try:
#         new_token = refresh_ig_token(IG_ACCESS_TOKEN)
#         update_github_secret("IG_ACCESS_TOKEN", new_token)
#         state["ig_last_refreshed"] = datetime.now(timezone.utc).isoformat()
#         save_json_state(drive, DRIVE_FOLDER_ID, TOKEN_STATE_FILENAME, state, state_file_id)
#         log("IG token refreshed successfully and GitHub secret updated.")
#         return new_token
#     except Exception as e:
#         log(f"IG TOKEN REFRESH FAILED: {e}")
#         send_alert(
#             "IG token refresh FAILED",
#             f"The automated Instagram token refresh failed: {e}\n\n"
#             f"The bot will keep using the existing token for now, but it may "
#             f"expire soon. Please refresh IG_ACCESS_TOKEN manually if this "
#             f"keeps happening.",
#         )
#         return IG_ACCESS_TOKEN


# # ---------------------------------------------------------------------------
# # Download + pad + caption extraction (from the link scripts, merged in)
# # ---------------------------------------------------------------------------

# def get_video_dimensions(path):
#     try:
#         result = subprocess.run(
#             ["ffprobe", "-v", "error", "-select_streams", "v:0",
#              "-show_entries", "stream=width,height", "-of", "json", str(path)],
#             capture_output=True, text=True, check=True,
#         )
#         data = json.loads(result.stdout)
#         stream = data["streams"][0]
#         return int(stream["width"]), int(stream["height"])
#     except Exception:
#         return None


# def pad_to_vertical(path, target_w=PAD_TARGET_W, target_h=PAD_TARGET_H):
#     """Scale+pad to target_w x target_h if not already that ratio. No-op if
#     already 9:16 (within rounding tolerance)."""
#     dims = get_video_dimensions(path)
#     if dims is None:
#         return False, "could not read video dimensions"

#     width, height = dims
#     target_ratio = target_w / target_h
#     current_ratio = width / height

#     if abs(current_ratio - target_ratio) < 0.01:
#         return True, "already 9:16, skipped"

#     tmp_path = path.with_name(path.stem + "_padded_tmp" + path.suffix) if hasattr(path, "with_name") else f"{path}.padded_tmp.mp4"

#     vf = (
#         f"scale={target_w}:{target_h}:force_original_aspect_ratio=decrease,"
#         f"pad={target_w}:{target_h}:(ow-iw)/2:(oh-ih)/2:color=black"
#     )
#     cmd = ["ffmpeg", "-y", "-i", str(path), "-vf", vf,
#            "-c:v", "libx264", "-preset", "medium", "-crf", "18",
#            "-c:a", "copy", str(tmp_path)]

#     try:
#         subprocess.run(cmd, capture_output=True, text=True, check=True)
#         os.replace(tmp_path, path)
#         return True, "padded to 9:16"
#     except subprocess.CalledProcessError as e:
#         if os.path.exists(tmp_path):
#             os.remove(tmp_path)
#         return False, f"ffmpeg error: {e.stderr[-300:] if e.stderr else e}"


# def download_and_prepare_video(link, dest_path):
#     """Downloads `link` to dest_path (e.g. '/tmp/xamo_poster/row5.mp4'),
#     pads it to 9:16 if needed, and returns (success, caption_or_None,
#     error_message_or_None)."""
#     ydl_opts = {
#         "outtmpl": dest_path,
#         "format": "bv*+ba/b",
#         "merge_output_format": "mp4",
#         "quiet": True,
#         "no_warnings": True,
#         "retries": 3,
#         "ignoreerrors": False,
#     }
#     try:
#         with yt_dlp.YoutubeDL(ydl_opts) as ydl:
#             info = ydl.extract_info(link, download=True)
#         caption = (info.get("description") or "").strip() if info else ""
#     except Exception as e:
#         return False, None, str(e)

#     if not os.path.exists(dest_path):
#         return False, None, "download reported success but output file is missing"

#     pad_ok, pad_msg = pad_to_vertical(dest_path)
#     if not pad_ok:
#         return False, None, f"9:16 padding failed: {pad_msg}"

#     return True, caption, None


# # ---------------------------------------------------------------------------
# # Instagram posting
# # ---------------------------------------------------------------------------

# def upload_temp_github_release_asset(local_video_path, filename):
#     headers = {"Authorization": f"Bearer {GH_PAT}", "Accept": "application/vnd.github+json"}

#     releases_resp = requests.get(
#         f"https://api.github.com/repos/{GH_REPO}/releases/tags/video-hosting",
#         headers=headers, timeout=20,
#     )
#     if releases_resp.status_code == 200:
#         release = releases_resp.json()
#     else:
#         create_resp = requests.post(
#             f"https://api.github.com/repos/{GH_REPO}/releases",
#             headers=headers,
#             json={"tag_name": "video-hosting", "name": "Temporary video hosting", "draft": False, "prerelease": True},
#             timeout=20,
#         )
#         release = create_resp.json()

#     upload_url = release["upload_url"].split("{")[0]

#     for existing_asset in release.get("assets", []):
#         if existing_asset.get("name") == filename:
#             log(f"IG: found stale leftover asset '{filename}' from a previous run, deleting it first.")
#             requests.delete(
#                 f"https://api.github.com/repos/{GH_REPO}/releases/assets/{existing_asset['id']}",
#                 headers=headers, timeout=20,
#             )

#     with open(local_video_path, "rb") as f:
#         upload_resp = requests.post(
#             f"{upload_url}?name={filename}",
#             headers={**headers, "Content-Type": "video/mp4"},
#             data=f,
#             timeout=120,
#         )
#     upload_resp.raise_for_status()
#     asset = upload_resp.json()
#     return asset["browser_download_url"], asset["id"]


# def delete_github_release_asset(asset_id):
#     headers = {"Authorization": f"Bearer {GH_PAT}", "Accept": "application/vnd.github+json"}
#     requests.delete(
#         f"https://api.github.com/repos/{GH_REPO}/releases/assets/{asset_id}",
#         headers=headers, timeout=20,
#     )


# def post_to_instagram(video_local_path, caption, access_token):
#     base_url = "https://graph.instagram.com"
#     video_url, asset_id = upload_temp_github_release_asset(video_local_path, os.path.basename(video_local_path))
#     log(f"IG: temporarily hosted video at {video_url}")

#     try:
#         create_resp = requests.post(
#             f"{base_url}/{IG_USER_ID}/media",
#             data={"media_type": "REELS", "video_url": video_url, "caption": caption, "access_token": access_token},
#         ).json()

#         if "id" not in create_resp:
#             log(f"IG: failed to create container: {create_resp}")
#             return False

#         container_id = create_resp["id"]
#         status = None
#         for _ in range(30):
#             time.sleep(10)
#             status_resp = requests.get(
#                 f"{base_url}/{container_id}",
#                 params={"fields": "status_code", "access_token": access_token},
#             ).json()
#             status = status_resp.get("status_code")
#             if status == "FINISHED":
#                 break
#             if status == "ERROR":
#                 log(f"IG: container processing failed: {status_resp}")
#                 return False

#         if status != "FINISHED":
#             log(f"IG: container did not finish processing in time (last status: {status})")
#             return False

#         publish_resp = requests.post(
#             f"{base_url}/{IG_USER_ID}/media_publish",
#             data={"creation_id": container_id, "access_token": access_token},
#         ).json()

#         if "id" not in publish_resp:
#             log(f"IG: publish failed: {publish_resp}")
#             return False

#         log(f"IG: published successfully, media id {publish_resp['id']}")
#         return True
#     finally:
#         delete_github_release_asset(asset_id)
#         log("IG: cleaned up temporary hosted video.")


# # ---------------------------------------------------------------------------
# # Facebook Page posting
# # ---------------------------------------------------------------------------

# def post_to_facebook_page(video_local_path, caption):
#     base_url = f"https://graph-video.facebook.com/{GRAPH_API_VERSION}"
#     with open(video_local_path, "rb") as f:
#         files = {"source": f}
#         data = {"description": caption, "access_token": FB_PAGE_ACCESS_TOKEN}
#         resp = requests.post(f"{base_url}/{FB_PAGE_ID}/videos", data=data, files=files).json()

#     if "id" not in resp:
#         log(f"FB: post failed: {resp}")
#         return False

#     log(f"FB: posted successfully, video id {resp['id']}")
#     return True


# # ---------------------------------------------------------------------------
# # YouTube posting
# # ---------------------------------------------------------------------------

# def get_youtube_service():
#     creds = UserCredentials(
#         token=None, refresh_token=YT_REFRESH_TOKEN,
#         client_id=YT_CLIENT_ID, client_secret=YT_CLIENT_SECRET,
#         token_uri="https://oauth2.googleapis.com/token",
#         scopes=["https://www.googleapis.com/auth/youtube.upload"],
#     )
#     return build("youtube", "v3", credentials=creds)


# def post_to_youtube(video_local_path, yt_title, description, publish_at_iso):
#     try:
#         youtube = get_youtube_service()
#         body = {
#             "snippet": {"title": yt_title[:100], "description": description, "categoryId": "22"},
#             "status": {"privacyStatus": "private", "publishAt": publish_at_iso, "selfDeclaredMadeForKids": False},
#         }
#         media = MediaFileUpload(video_local_path, chunksize=-1, resumable=True, mimetype="video/mp4")
#         request = youtube.videos().insert(part="snippet,status", body=body, media_body=media)
#         response = None
#         while response is None:
#             status, response = request.next_chunk()
#         log(f"YT: uploaded successfully, video id {response['id']}, scheduled for {publish_at_iso}")
#         return True
#     except Exception as e:
#         log(f"YT: upload failed: {e}")
#         return False


# # ---------------------------------------------------------------------------
# # Healthcheck ping
# # ---------------------------------------------------------------------------

# def ping_healthcheck(success=True, message=""):
#     try:
#         url = HEALTHCHECK_PING_URL if success else f"{HEALTHCHECK_PING_URL}/fail"
#         requests.post(url, data=message.encode("utf-8"), timeout=10)
#     except Exception as e:
#         log(f"Healthcheck ping failed (non-fatal): {e}")


# # ---------------------------------------------------------------------------
# # Per-platform attempt helper (Download / IG / FB: retry counting + skip-after-N)
# # ---------------------------------------------------------------------------

# def attempt_with_retries(ws, row_num, row, status_col, fails_col, posted_at_col,
#                           label, post_fn, on_skipped_error=None):
#     """Generic 3-strike attempt helper used for Download, IG, and FB.
#     post_fn() must return True/False (success/failure)."""
#     if row[status_col] in DONE_STATES:
#         return False

#     log(f"Attempting {label} for row {row_num}...")
#     success = post_fn()

#     if success:
#         set_cell(ws, row_num, status_col, "downloaded" if status_col == "DL_Status" else "posted")
#         if posted_at_col:
#             set_cell(ws, row_num, posted_at_col, datetime.now(timezone.utc).isoformat())
#         set_cell(ws, row_num, fails_col, 0)
#     else:
#         fails = (row[fails_col] or 0) + 1
#         set_cell(ws, row_num, fails_col, fails)
#         if fails >= MAX_FAILURES_PER_PLATFORM:
#             set_cell(ws, row_num, status_col, "skipped_error")
#             log(f"{label}: hit {fails} failures on row {row_num}, marking skipped_error.")
#             send_alert(
#                 f"{label} skipped after {fails} failures",
#                 f"Row {row_num} (link: {row['Link']}) failed {label} {fails} times in a row "
#                 f"and has been marked 'skipped_error'. It will NOT be retried automatically. "
#                 f"Check the row in queue.xlsx and fix/reset it manually.",
#             )
#             if on_skipped_error:
#                 on_skipped_error()
#         else:
#             set_cell(ws, row_num, status_col, "failed")

#     return True


# def attempt_youtube_no_retry(ws, row_num, row, post_fn):
#     if row["YT_Status"] in DONE_STATES:
#         return False

#     log(f"Attempting YouTube post for row {row_num}...")
#     success = post_fn()

#     if success:
#         set_cell(ws, row_num, "YT_Status", "posted")
#         set_cell(ws, row_num, "YT_PostedAt", datetime.now(timezone.utc).isoformat())
#     else:
#         set_cell(ws, row_num, "YT_Status", "failed")
#         log(f"YouTube: row {row_num} failed this cycle -- will not be retried (no strike-counter for YT).")

#     return True


# # ---------------------------------------------------------------------------
# # Main
# # ---------------------------------------------------------------------------

# def run():
#     log("Starting poster run.")
#     drive = get_drive_service()

#     ig_access_token = maybe_refresh_ig_token(drive)

#     queue_file_id = find_file_in_folder(drive, DRIVE_FOLDER_ID, QUEUE_FILENAME)
#     if not queue_file_id:
#         log(f"ERROR: could not find {QUEUE_FILENAME} in the Drive folder.")
#         ping_healthcheck(success=False, message="queue.xlsx not found")
#         sys.exit(1)

#     local_queue_path = os.path.join(LOCAL_WORKDIR, QUEUE_FILENAME)
#     download_file(drive, queue_file_id, local_queue_path)
#     wb, ws = load_queue(local_queue_path)

#     run_state, run_state_file_id = load_json_state(
#         drive, DRIVE_FOLDER_ID, RUN_STATE_FILENAME, {"successful_run_count": 0}
#     )

#     ig_fb_posted_this_run = False
#     row_num_for_ig_fb = None  # the row IG/FB actually succeeded on this run, if any

#     # =======================================================================
#     # Find current row and (if needed) advance through consecutive dead
#     # links within THIS run, until one downloads successfully or we run
#     # out of rows.
#     # =======================================================================
#     downloaded_row_num = None
#     downloaded_local_path = None
#     downloaded_caption = None

#     for row_num, row in all_data_rows(ws):
#         # Skip rows that are entirely done already.
#         if row["IG_Status"] in DONE_STATES and row["FB_Status"] in DONE_STATES:
#             continue
#         # Skip rows permanently dead on download.
#         if row["DL_Status"] == "skipped_error":
#             continue

#         # If this row hasn't downloaded successfully yet, try it now.
#         if row["DL_Status"] != "downloaded":
#             local_video_path = os.path.join(LOCAL_WORKDIR, f"row{row_num}.mp4")

#             def do_download():
#                 ok, caption, err = download_and_prepare_video(row["Link"], local_video_path)
#                 if not ok:
#                     log(f"Download failed for row {row_num}: {err}")
#                     return False
#                 nonlocal downloaded_caption
#                 downloaded_caption = caption
#                 return True

#             attempted = attempt_with_retries(
#                 ws, row_num, row, "DL_Status", "DL_Fails", None,
#                 "Download", do_download,
#             )

#             row_after = get_row_dict(ws, row_num)
#             if row_after["DL_Status"] == "downloaded":
#                 downloaded_row_num = row_num
#                 downloaded_local_path = local_video_path
#                 resolved_caption = downloaded_caption if downloaded_caption else DEFAULT_CAPTION
#                 set_cell(ws, row_num, "Caption", resolved_caption)
#                 break  # found our row for this run
#             else:
#                 # Either "failed" (will retry next run naturally since this
#                 # loop always starts from the top) or just hit
#                 # "skipped_error" this attempt -- either way, move on to
#                 # the NEXT row within this same run, per spec.
#                 if row_after["DL_Status"] == "failed":
#                     # Still within its 3 strikes -- don't treat this as the
#                     # row for this run, but also don't advance past it for
#                     # FUTURE runs (it stays first-in-line next time). We do
#                     # still need to try the next row THIS run though.
#                     continue
#                 else:
#                     continue
#         else:
#             # Already downloaded in a previous run but IG/FB not fully done
#             # yet -- this is our row for this run, no need to re-download.
#             downloaded_row_num = row_num
#             downloaded_local_path = os.path.join(LOCAL_WORKDIR, f"row{row_num}.mp4")
#             # Re-download since the local disk is ephemeral between runs.
#             row_link = row["Link"]

#             def do_redownload():
#                 ok, caption, err = download_and_prepare_video(row_link, downloaded_local_path)
#                 if not ok:
#                     log(f"Re-download failed for already-downloaded row {row_num}: {err}")
#                 return ok

#             if not do_redownload():
#                 log(f"ERROR: could not re-fetch previously-downloaded video for row {row_num}. Skipping this run.")
#                 downloaded_row_num = None
#                 downloaded_local_path = None
#                 continue
#             break

#     if downloaded_row_num is None:
#         log("No row available to process this run (nothing new, or all remaining rows are exhausted/skipped).")
#     else:
#         row = get_row_dict(ws, downloaded_row_num)
#         caption = row["Caption"] or DEFAULT_CAPTION
#         log(f"Row for this run: {downloaded_row_num}")

#         ig_ok = attempt_with_retries(
#             ws, downloaded_row_num, row, "IG_Status", "IG_Fails", "IG_PostedAt",
#             "Instagram", lambda: post_to_instagram(downloaded_local_path, caption, ig_access_token),
#         )
#         row = get_row_dict(ws, downloaded_row_num)
#         fb_ok = attempt_with_retries(
#             ws, downloaded_row_num, row, "FB_Status", "FB_Fails", "FB_PostedAt",
#             "Facebook", lambda: post_to_facebook_page(downloaded_local_path, caption),
#         )
#         row = get_row_dict(ws, downloaded_row_num)

#         if row["IG_Status"] == "posted" and row["FB_Status"] == "posted":
#             ig_fb_posted_this_run = True
#             row_num_for_ig_fb = downloaded_row_num

#     # =======================================================================
#     # Run counter + YouTube (every Nth successful run)
#     # =======================================================================
#     if ig_fb_posted_this_run:
#         run_state["successful_run_count"] = run_state.get("successful_run_count", 0) + 1
#         count = run_state["successful_run_count"]
#         log(f"IG+FB succeeded this run. Successful run count is now {count}.")

#         if count % YT_EVERY_N_RUNS == 0:
#             log(f"This is run #{count} -- YouTube's turn (every {YT_EVERY_N_RUNS} successful runs).")
#             row = get_row_dict(ws, row_num_for_ig_fb)
#             caption = row["Caption"] or DEFAULT_CAPTION
#             yt_title = (caption[:90] + " #shorts") if caption else "#shorts"
#             now_iso = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

#             attempt_youtube_no_retry(
#                 ws, row_num_for_ig_fb, row,
#                 lambda: post_to_youtube(downloaded_local_path, yt_title, caption, now_iso),
#             )
#         else:
#             log(f"Not YouTube's turn this cycle ({count % YT_EVERY_N_RUNS} of {YT_EVERY_N_RUNS}).")

#         save_json_state(drive, DRIVE_FOLDER_ID, RUN_STATE_FILENAME, run_state, run_state_file_id)
#     else:
#         log("IG+FB did not both succeed this run -- run counter not incremented, YouTube not considered.")

#     wb.save(local_queue_path)
#     upload_or_replace_file(drive, DRIVE_FOLDER_ID, QUEUE_FILENAME, local_queue_path, existing_file_id=queue_file_id)

#     if downloaded_row_num is not None:
#         ping_healthcheck(success=True, message="Run completed with activity")
#     else:
#         ping_healthcheck(success=True, message="Nothing needed posting this run")

#     log("Run complete.")


# def main():
#     try:
#         run()
#     except Exception as e:
#         log(f"FATAL ERROR in run(): {e}")
#         ping_healthcheck(success=False, message=f"Fatal error: {e}")
#         try:
#             send_alert("Xamo Auto Poster run CRASHED", f"The bot run failed with an unhandled error:\n\n{e}")
#         except Exception:
#             log("Could not send crash alert either.")
#         sys.exit(1)


# if __name__ == "__main__":
#     main()
    
